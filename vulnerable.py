import os
import re
import sys
import html
import time
import queue
import zipfile
import datetime
import openpyxl
import Coredata.load


class Vul_re(object):
    def __init__(self):
        super(Vul_re, self).__init__()
        self.vul_list_re = '<python>host<python>.*?<td valign="top".*?<th width="120">IP地址</th>.*?<td>(.*?)</td>.*?</td>.*?<python>host</python>.*?<python>vul_list<python>(.*?)<python>vul_list</python>'
        self.vul_detail_re = '<python>vul_detail<python>(.*?)<python>vul_detail</python>'
        self.vul_details_re = '<python>vul_details<python>(.*?)<python>vul_details</python>'

        self.danger_re = '<span class="level_danger_(.*?)".*?table_\d_(\d+).*?>(.*?)</span>'
        self.title_re = '<python>title<python>(.*?)<python>title</python>'
        self.other_re = '<td class="vul_port">(.*?)</td>.*?<td>(.*?)</td>.*?<td>(.*?)</td>.*?<td>.*?<ul>(.*?)</ul>'

class Vul_content(object):
    def __init__(self,vul_re):
        super(Vul_content, self).__init__()
        self.vul_list_content = re.findall(vul_re.vul_list_re,htmlcont,re.S|re.M)
        self.vul_detail_content = re.findall(vul_re.vul_detail_re,htmlcont,re.S|re.M)

class Solve_re(object):
    def __init__(self):
        super(Solve_re, self).__init__()
        self.solve_re = '<th width="100">解决办法</th>.*?<td>(.*?)</td>'
        self.describe_re = '<tr class="solution.*?table_\d_(\d+).*?<th width="100">详细描述</th>.*?<td>(.*?)</td>'
        self.cve_re = '<th width="100">CVE编号</th>.*?<td><a target=.*?>(.*?)</a>.*?</td>'


class Other(object):
    def __init__(self, vul_re, all_vuln_list):
        super(Other, self).__init__()
        self.all_other = re.findall(vul_re.other_re,all_vuln_list,re.S|re.M)

class Danger(object):
    def __init__(self, vul_re, other):
        super(Danger, self).__init__()
        self.danger_coneent = re.findall(vul_re.danger_re,other,re.S|re.M)

class Solve(object):
    def __init__(self, solve, all_vul_details):
        super(Solve, self).__init__()
        self.solve_plumb = re.findall(solve.solve_re,all_vul_details,re.S|re.M)
        self.describe_plumb = re.findall(solve.describe_re,all_vul_details,re.S|re.M)
        self.cve_plumb = re.findall(solve.cve_re,all_vul_details,re.S|re.M)

def main():
    folder_name = str(sys.argv[1])
    print('\n数据提取完成，正在生成漏洞跟踪表...')
    wb = openpyxl.Workbook()
    ws = wb.active

    vul_re = Vul_re()

    with open('./Coredata/database.mdb') as content:
        for zip_content in content:
            zip_cont = zip_content.strip('\n\r')
            content = open(zip_cont,'r')
            global htmlcont
            htmlcont = content.read()
            content.close()
            #正则：<python>title<python>(.*?)<python>title</python>
            sheet_name =  re.findall(vul_re.title_re,htmlcont,re.S|re.M)
            print('正在导出 %s'%sheet_name[0])
            ws = wb.create_sheet(sheet_name[0],0)
            ws['A1'] = '序号'
            ws['B1'] = '主机名'
            ws['C1'] = 'IP地址'
            ws['D1'] = '漏洞名称'
            ws['E1'] = '风险分类'
            ws['F1'] = '风险等级'
            ws['G1'] = '整改建议'
            ws['H1'] = '漏洞描述'
            ws['I1'] = '漏洞CVE编号'
            ws['J1'] = '漏洞对应端口'
            ws['K1'] = '漏洞对应协议'
            ws['L1'] = '漏洞对应服务'

            #正则，获取漏洞概况、漏洞详情
            vul_content = Vul_content(vul_re)

            vul_all_list = []
            vul_all_detail = []

            for all_vul_list in vul_content.vul_list_content:
                for other in Other(vul_re,all_vul_list[1]).all_other:
                    for danger in Danger(vul_re,other[3]).danger_coneent:
                        vul_all_list.append([danger[1],all_vul_list[0],danger[2],danger[0].replace('low','低').replace('middle','中').replace('high','高'),other[0],other[1],other[2]])

            for all_vul_detail in vul_content.vul_detail_content:
                vul_details_content = re.findall(vul_re.vul_details_re,all_vul_detail,re.S|re.M)
                for all_vul_details in vul_details_content:
                    vul_detail = Solve(Solve_re(),all_vul_details)
                    # #迭代
                    for solve,describe in zip(vul_detail.solve_plumb,vul_detail.describe_plumb):
                        cve = vul_detail.cve_plumb
                        if cve:
                            pass
                        else:
                            cve = ['漏洞暂无CVE编号']
                        vul_all_detail.append([describe[0],html.unescape(re.sub('<br/>','',solve)).replace(' ','').strip('\n'),html.unescape(re.sub('<br/>','',describe[1])).replace(' ','').strip('\n'),cve[0]])

            for line_vul_list in vul_all_list:
                vul_list.put(line_vul_list)

            i = 1
            while not vul_list.empty():
                wait_list = vul_list.get()
                for wait_detail in vul_all_detail:
                    if wait_list[0] == wait_detail[0]:
                        ws.append([i,'',wait_list[1],wait_list[2],'漏洞',wait_list[3],wait_detail[1],wait_detail[2],wait_detail[3],wait_list[4],wait_list[5],wait_list[6]])
                        i += 1
                        break

    wb.save(folder_name+'/漏洞跟踪表.xlsx')
    print('\n漏洞跟踪表导出完成，保存在 %s 目录下。'%folder_name)




if __name__ == '__main__':
    starttime = datetime.datetime.now()
    Coredata.load.start_date()
    Coredata.load.load_date()
    breaktime = datetime.datetime.now()
    print ('数据提取花时：%s秒...'%(breaktime - starttime).seconds)
    vul_list = queue.Queue()
    main()
    Coredata.load.end_date()
    endtime = datetime.datetime.now()
    print ('导出花时：%s秒...'%(endtime - breaktime).seconds)
